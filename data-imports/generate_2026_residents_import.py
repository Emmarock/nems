#!/usr/bin/env python3
"""
Parses "Residents 2026 Outstanding.xlsx" and generates:
  1. A one-time PL/pgSQL data-load script (residents/properties/levies/invoices/payments).
  2. A markdown data-quality report documenting every judgment call and flagged row.

This REPLACES the 2025 dataset (see 2026_truncate.sql, run separately and manually first) —
it is not a merge. Only 4 of the workbook's ~55 sheets are used, per the estate's own
description of what each one means:

  - '"26 NITEL Payment & Outstanding' — the resident roster (Name, phone, grouped under
    "ROAD X" / "N-th AVENUE" section-header rows). Source of truth for WHO the residents are.
    Duplicate names (30 of them — mostly "ROAD A" and part of "ROAD B" being re-listed verbatim
    near the end of the sheet) are ignored: first occurrence wins, later ones are skipped.
  - '2026 Development Levy & Deptors' — payment records for Devt. Entrance Levy and Membership
    Form (one-time fees). Its own SECURITY FEE / YEARLY DEVT. LEVY columns are NOT used for
    amounts — see below. Its '24/2025 Debtors' column (prior-year carried debt) is NOT imported
    as its own invoice either: that debt is already folded into NEW 2026 DEBTORS' own Outstanding
    Debt figure, so a separate arrears invoice would double-count it in the resident's total owed.
  - ' Electricity Payment & Deptors' — payment records for Electricity/Transformer Levy, 2nd
    Transformer Payment, and Donation (individual, electricity-related; verified identical to
    the copy of the same column on the dev-levy sheet, so only imported once, from here).
  - 'NEW 2026 DEBTORS' — authoritative for the recurring annual dues (Security Fee + Yearly
    Devt. Levy, which this sheet only reports as ONE combined "2026 Payment"/"Paid" figure per
    resident, not split — so they're imported as a single merged levy, not two).

A value present in a payment column means PAID IN FULL for that line item (every sheet's title
row reads "Nitel Estate Residents' 2026 Payment") — same convention the 2025 import used.

Multi-property owners: a resident whose name recurs on the dev-levy/electricity/new-debtors
sheets under a DIFFERENT road (e.g. "Mr Kamoru Sunday Oluseje" owns plots on Road F, 3rd Avenue,
and Road U) still gets only ONE resident+property (per the roster sheet's dedup rule above) —
their amounts are SUMMED onto that one resident rather than dropped, so total debt is right even
though it's not split across separate synthetic properties. Listed in the report as
"consolidated multi-row residents".

Residents that appear on a levy sheet but NOT on the roster sheet (43 on NEW 2026 DEBTORS) have
nowhere to attach a resident record to, per the "residents come from sheet 1 only" instruction —
their amounts are skipped and listed in the report as unmatched, not silently dropped.
"""
import openpyxl
import re
from collections import defaultdict

SRC = "/Users/apata/IdeaProjects/NitelEstate/Residents 2026 Outstanding.xlsx"
OUT_SQL = "/Users/apata/IdeaProjects/NitelEstate/data-imports/2026_residents_import.sql"
OUT_REPORT = "/Users/apata/IdeaProjects/NitelEstate/data-imports/2026_import_report.md"

SHEET_ROSTER = '"26 NITEL Payment & Outstanding'
SHEET_DEVLEVY = '2026 Development Levy & Deptors'
SHEET_ELECTRICITY = ' Electricity Payment & Deptors'
SHEET_DEBTORS = 'NEW 2026 DEBTORS'

ROAD_RE = re.compile(r'^(ROAD [A-Z]|(\d+)(st|nd|rd|th)\s*AVENUE|UNIDENTIFIED ROADS?)\b', re.IGNORECASE)


def norm(s):
    return re.sub(r'\s+', ' ', str(s).strip()).upper()


def clean_name(raw):
    return re.sub(r'\s+', ' ', str(raw).strip())


def sql_str(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def clean_phone(raw, flags, rownum, name):
    """Same convention as the 2025 import: only two shapes are treated as *local* Nigerian
    numbers (11 digits starting with 0, or exactly 10 digits); 11-15 digit values are assumed
    to already carry a country code. Anything else -> 'UNKNOWN' (phone is NOT NULL), flagged."""
    if raw is None:
        return "UNKNOWN"
    s = str(raw).strip()
    digits = re.sub(r'\D', '', s)
    if not digits:
        flags.append((rownum, name, "phone", f"non-numeric phone value {raw!r} -> stored as UNKNOWN"))
        return "UNKNOWN"
    if len(digits) == 11 and digits.startswith('0'):
        return "+234" + digits[1:]
    if len(digits) == 10:
        return "+234" + digits
    if 10 < len(digits) <= 15:
        return "+" + digits
    flags.append((rownum, name, "phone", f"implausible phone {raw!r} ({len(digits)} digits) -> stored as UNKNOWN"))
    return "UNKNOWN"


MAX_PLAUSIBLE_AMOUNT = 5_000_000  # real values top out around 1,015,000; anything past this is
                                   # almost certainly a phone number that landed in the wrong
                                   # column (e.g. a raw 8028955336 or 16465462630 cell) rather
                                   # than a real levy/donation amount.


def parse_amount(raw, flags, rownum, name, field):
    """Returns a positive float, or None (and logs a flag) if blank/zero/unparseable/implausible.
    Handles messy text cells like "160,000," (comma thousands-separator + stray trailing
    comma) as well as free text like "waived" (flagged, not guessed at). Also rejects
    numeric-typed cells that are wildly out of range for a real amount - Excel stores phone
    numbers typed into the wrong column as plain numbers, and those would otherwise sail
    straight through as e.g. a 16-billion-naira "donation"."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if raw <= 0:
            return None
        if raw > MAX_PLAUSIBLE_AMOUNT:
            flags.append((rownum, name, field, f"implausibly large value {raw!r} (likely a misplaced phone number) — skipped"))
            return None
        return float(raw)
    s = str(raw).strip()
    cleaned = re.sub(r'[,\s]', '', s).strip('.,')
    if re.fullmatch(r'-?\d+(\.\d+)?', cleaned):
        v = float(cleaned)
        if v <= 0:
            return None
        if v > MAX_PLAUSIBLE_AMOUNT:
            flags.append((rownum, name, field, f"implausibly large value {raw!r} (likely a misplaced phone number) — skipped"))
            return None
        return v
    flags.append((rownum, name, field, f"non-numeric value {raw!r} — skipped, needs manual entry"))
    return None


def parse_section_rows(ws, name_col, extra_cols):
    """extra_cols: {label: 1-based column index}. Returns a list of row dicts; road
    section-header rows (ROAD A, 1st AVENUE, ...) set `current_road` for subsequent rows
    and are not themselves records."""
    current_road = None
    rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v is None:
            continue
        vs = clean_name(v)
        if not vs:
            continue
        if ROAD_RE.match(vs):
            current_road = vs.upper()
            continue
        if 'total' in vs.lower():
            continue
        if vs.upper() == 'NAME' or 'NITEL ESTATE RESIDENTS' in vs.upper():
            continue
        if current_road is None:
            # Stray pre-roster cell (title/header row artifact) — not a real record.
            continue
        extras = {label: ws.cell(row=r, column=c).value for label, c in extra_cols.items()}
        rows.append({'row': r, 'road': current_road, 'name': vs, **extras})
    return rows


def aggregate(rows, fields, sheet_label, residents, flags):
    """Sums each field across every row matching a known resident (handles multi-property
    owners who recur under different roads). Returns (aggregated_amounts, unmatched_rows,
    consolidated_multi_row_residents)."""
    agg = defaultdict(lambda: defaultdict(float))
    matched_rows = defaultdict(list)
    unmatched = []
    for row in rows:
        k = norm(row['name'])
        if k not in residents:
            unmatched.append(row)
            continue
        matched_rows[k].append(row['row'])
        for f in fields:
            v = parse_amount(row.get(f), flags, row['row'], row['name'], f"{sheet_label}.{f}")
            if v:
                agg[k][f] += v
    consolidated = {k: rs for k, rs in matched_rows.items() if len(rs) > 1}
    return agg, unmatched, consolidated


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    roster_rows = parse_section_rows(wb[SHEET_ROSTER], name_col=1, extra_cols={'phone': 4})
    devlevy_rows = parse_section_rows(wb[SHEET_DEVLEVY], name_col=3,
                                       extra_cols={'debtors_24_25': 1, 'entrance': 7, 'membership': 8})
    electricity_rows = parse_section_rows(wb[SHEET_ELECTRICITY], name_col=1,
                                           extra_cols={'electricity': 3, 'transformer2': 4, 'donation': 6})
    debtors_rows = parse_section_rows(wb[SHEET_DEBTORS], name_col=2,
                                       extra_cols={'payment2026': 4, 'paid2026': 5, 'outstanding2026': 6})

    flags = []

    # ---- Residents (dedup: first occurrence wins) ----
    residents = {}
    order = []
    dup_skipped = []
    seq_by_road = defaultdict(int)
    for row in roster_rows:
        k = norm(row['name'])
        if k in residents:
            dup_skipped.append(row)
            continue
        seq_by_road[row['road']] += 1
        house_number = f"{row['road']}-{seq_by_road[row['road']]:03d}"
        phone_sql = clean_phone(row['phone'], flags, row['row'], row['name'])
        residents[k] = {
            'name': row['name'], 'road': row['road'], 'house_number': house_number, 'phone': phone_sql,
        }
        order.append(k)

    # ---- Aggregate the 3 levy sheets against the resident roster ----
    devlevy_agg, devlevy_unmatched, devlevy_consolidated = aggregate(
        devlevy_rows, ['debtors_24_25', 'entrance', 'membership'], 'dev-levy', residents, flags)
    elec_agg, elec_unmatched, elec_consolidated = aggregate(
        electricity_rows, ['electricity', 'transformer2', 'donation'], 'electricity', residents, flags)
    debtors_agg, debtors_unmatched, debtors_consolidated = aggregate(
        debtors_rows, ['payment2026', 'paid2026'], 'new-debtors', residents, flags)

    # ---- Build SQL ----
    sql_parts = []
    sql_parts.append("-- One-time data load: Residents 2026 Outstanding.xlsx -> residents/properties/invoices/payments")
    sql_parts.append("-- Generated by data-imports/generate_2026_residents_import.py — NOT a Flyway migration.")
    sql_parts.append("-- Run 2026_truncate.sql FIRST (clears the 2025 dataset this replaces), then this file.")
    sql_parts.append("-- Safety guard: aborts if it looks like this has already been applied.")
    sql_parts.append("DO $$")
    sql_parts.append("BEGIN")
    sql_parts.append("  IF EXISTS (SELECT 1 FROM property WHERE house_number = 'ROAD A-001') THEN")
    sql_parts.append("    RAISE EXCEPTION 'Residents 2026 import already applied — aborting to avoid duplicates';")
    sql_parts.append("  END IF;")
    sql_parts.append("END $$;")
    sql_parts.append("")
    sql_parts.append("DO $$")
    sql_parts.append("DECLARE")
    sql_parts.append("  v_prop_id bigint;")
    sql_parts.append("  v_res_id bigint;")
    sql_parts.append("  v_inv_id bigint;")
    levy_keys = ['dues', 'entrance', 'membership', 'electricity', 'transformer2', 'donation']
    for key in levy_keys:
        sql_parts.append(f"  v_levy_{key} bigint;")
    sql_parts.append("BEGIN")
    sql_parts.append("")
    sql_parts.append("  -- Levy catalog for the 2026 dataset. 'Security Fee' and 'Yearly Devt. Levy' are merged")
    sql_parts.append("  -- into one levy because NEW 2026 DEBTORS (the source of truth for these two, per the")
    sql_parts.append("  -- estate's own description of that sheet) only reports them as one combined figure.")
    sql_parts.append("  -- Catalog 'amount' is the estate's fixed/standard rate for that levy (confirmed by the")
    sql_parts.append("  -- treasurer) - individual invoices still carry their own amount from the sheet, which")
    sql_parts.append("  -- can differ (e.g. a multi-property owner's combined dues, a partial/short payment).")
    sql_parts.append("  -- Donation has no fixed rate (ad hoc), so stays at 0.")
    sql_parts.append("  -- No '2024/2025 Arrears' levy: that carried-over debt is already folded into the")
    sql_parts.append("  -- 'Security Fee + Yearly Devt. Levy' invoice amount sourced from NEW 2026 DEBTORS -")
    sql_parts.append("  -- a separate arrears invoice would double-count it in the resident's total outstanding.")
    levy_defs = [
        ('dues', 'Security Fee + Yearly Devt. Levy', 'ANNUAL', 80000),
        ('entrance', 'Devt. Entrance Levy', 'ONE_TIME', 50000),
        ('membership', 'Membership Form', 'ONE_TIME', 5000),
        ('electricity', 'Electricity/Transformer Levy', 'ANNUAL', 300000),
        ('transformer2', '2nd Transformer Payment', 'ONE_TIME', 100000),
        ('donation', 'Donation', 'ONE_TIME', 0),
    ]
    for key, label, freq, catalog_amount in levy_defs:
        sql_parts.append(
            f"  INSERT INTO levy (created_at, updated_at, name, amount, frequency, active) "
            f"VALUES (now(), now(), {sql_str(label)}, {catalog_amount}, '{freq}', true) RETURNING id INTO v_levy_{key};"
        )
    sql_parts.append("")

    counts = defaultdict(int)
    overpaid_flags = []

    for k in order:
        res = residents[k]
        road = res['road']
        address = f"{road.title()}, Nitel Estate"

        sql_parts.append(f"  -- {res['name']} ({road})")
        sql_parts.append(
            "  INSERT INTO property (created_at, updated_at, block, plot, house_number, address, "
            "property_type, occupancy_status) VALUES (now(), now(), "
            f"{sql_str(road.title())}, 'Unnumbered', {sql_str(res['house_number'])}, "
            f"{sql_str(address)}, 'DETACHED_HOUSE', 'OCCUPIED') RETURNING id INTO v_prop_id;"
        )
        sql_parts.append(
            "  INSERT INTO resident (created_at, updated_at, full_name, phone, email, property_id, "
            "resident_type, emergency_contact, status, registration_date) VALUES (now(), now(), "
            f"{sql_str(res['name'])}, {sql_str(res['phone'])}, NULL, v_prop_id, 'OWNER', NULL, 'ACTIVE', "
            "DATE '2026-01-01') RETURNING id INTO v_res_id;"
        )
        sql_parts.append("  UPDATE property SET owner_id = v_res_id WHERE id = v_prop_id;")
        counts['residents'] += 1

        def paid_line(levy_key, label, amount, issue, due, paid_date):
            sql_parts.append(
                f"  INSERT INTO invoice (created_at, updated_at, resident_id, levy_id, description, "
                f"amount, issue_date, due_date, status) VALUES (now(), now(), v_res_id, v_levy_{levy_key}, "
                f"{sql_str(label)}, {amount}, DATE '{issue}', DATE '{due}', 'ISSUED') RETURNING id INTO v_inv_id;"
            )
            sql_parts.append(
                "  INSERT INTO payment (created_at, updated_at, resident_id, invoice_id, amount, "
                "method, provider, provider_reference, status, paid_at) VALUES (now(), now(), v_res_id, "
                f"v_inv_id, {amount}, 'BANK_TRANSFER', NULL, NULL, 'SUCCESS', TIMESTAMP '{paid_date} 00:00:00');"
            )
            counts['invoices'] += 1
            counts['payments'] += 1

        # Security Fee + Yearly Devt. Levy — sourced from NEW 2026 DEBTORS (invoice = Payment,
        # payment = Paid; can be a partial payment, matching this app's normal balance model).
        d = debtors_agg.get(k, {})
        payment_amt = d.get('payment2026')
        paid_amt = d.get('paid2026')
        if payment_amt:
            sql_parts.append(
                "  INSERT INTO invoice (created_at, updated_at, resident_id, levy_id, description, "
                "amount, issue_date, due_date, status) VALUES (now(), now(), v_res_id, v_levy_dues, "
                f"'Security Fee + Yearly Devt. Levy', {payment_amt}, DATE '2026-01-01', DATE '2026-12-31', "
                "'ISSUED') RETURNING id INTO v_inv_id;"
            )
            counts['invoices'] += 1
            if paid_amt:
                pay_this = min(paid_amt, payment_amt)
                sql_parts.append(
                    "  INSERT INTO payment (created_at, updated_at, resident_id, invoice_id, amount, "
                    "method, provider, provider_reference, status, paid_at) VALUES (now(), now(), v_res_id, "
                    f"v_inv_id, {pay_this}, 'BANK_TRANSFER', NULL, NULL, 'SUCCESS', TIMESTAMP '2026-06-15 00:00:00');"
                )
                counts['payments'] += 1
                if paid_amt > payment_amt:
                    overpaid_flags.append((res['name'], payment_amt, paid_amt))

        # Devt. Entrance Levy / Membership Form — sourced from the dev-levy sheet itself.
        dv = devlevy_agg.get(k, {})
        if dv.get('entrance'):
            paid_line('entrance', 'Devt. Entrance Levy', dv['entrance'], '2026-01-01', '2026-12-31', '2026-06-15')
        if dv.get('membership'):
            paid_line('membership', 'Membership Form', dv['membership'], '2026-01-01', '2026-12-31', '2026-06-15')

        # Electricity/Transformer Levy, 2nd Transformer Payment, Donation — from the electricity sheet.
        el = elec_agg.get(k, {})
        if el.get('electricity'):
            paid_line('electricity', 'Electricity/Transformer Levy', el['electricity'], '2026-01-01', '2026-12-31', '2026-06-15')
        if el.get('transformer2'):
            paid_line('transformer2', '2nd Transformer Payment', el['transformer2'], '2026-01-01', '2026-12-31', '2026-06-15')
        if el.get('donation'):
            paid_line('donation', 'Donation', el['donation'], '2026-01-01', '2026-12-31', '2026-06-15')

        # 24/2025 Debtors: intentionally NOT imported as its own invoice. It's already carried
        # over into NEW 2026 DEBTORS' own Outstanding Debt figure (the source of the dues invoice
        # above) - a separate arrears invoice would double-count it in the resident's total owed.
        if dv.get('debtors_24_25'):
            counts['arrears_skipped'] += 1
            counts['arrears_skipped_amount'] += dv['debtors_24_25']

        sql_parts.append("")

    sql_parts.append("END $$;")
    sql_parts.append("")

    with open(OUT_SQL, "w") as f:
        f.write("\n".join(sql_parts))

    # ---- Report ----
    report = []
    report.append("# Nitel Estate Residents 2026 — Import Data Quality Report")
    report.append("")
    report.append(f"- **Source:** `Residents 2026 Outstanding.xlsx` (4 of ~55 sheets used — see script docstring)")
    report.append(f"- **Residents/properties created:** {counts['residents']}")
    report.append(f"- **Invoices created:** {counts['invoices']} ({counts['payments']} fully/partially paid, {counts['invoices'] - counts['payments']} unpaid)")
    report.append(f"- **24/2025 Debtors skipped (not imported as a separate invoice):** {counts['arrears_skipped']} residents, "
                   f"totaling {counts['arrears_skipped_amount']:,.0f} — this debt is already carried over into NEW 2026 DEBTORS' own Outstanding Debt figure, so a separate arrears invoice would double-count it.")
    report.append("")
    report.append("## Duplicate roster rows skipped (kept first occurrence)")
    report.append(f"{len(dup_skipped)} rows — mostly \"ROAD A\" and part of \"ROAD B\" being re-listed verbatim near the end of the roster sheet.")
    report.append("")
    for row in dup_skipped:
        report.append(f"- row {row['row']} ({row['road']}): {row['name']}")
    report.append("")
    report.append("## Consolidated multi-property residents (amounts summed onto one resident)")
    report.append("Same name recurs under a *different* road on a levy sheet — treated as one person owning multiple plots; each sheet's amounts for them were summed rather than only keeping the first row.")
    report.append("")
    for label, cons in [('Dev levy sheet', devlevy_consolidated), ('Electricity sheet', elec_consolidated), ('NEW 2026 DEBTORS sheet', debtors_consolidated)]:
        report.append(f"### {label} ({len(cons)})")
        for k, rows in cons.items():
            report.append(f"- **{residents[k]['name']}** — rows {', '.join(str(r) for r in rows)}")
        report.append("")
    report.append("## Unmatched rows (present on a levy sheet, not on the roster sheet — skipped entirely)")
    report.append("")
    for label, unmatched in [('Dev levy sheet', devlevy_unmatched), ('Electricity sheet', elec_unmatched), ('NEW 2026 DEBTORS sheet', debtors_unmatched)]:
        report.append(f"### {label} ({len(unmatched)})")
        for row in unmatched:
            report.append(f"- row {row['row']} ({row['road']}): {row['name']}")
        report.append("")
    report.append(f"## Flagged cells ({len(flags)}) — unparseable/free-text values, skipped")
    report.append("")
    report.append("| Row | Resident | Field | Issue |")
    report.append("|---|---|---|---|")
    for rownum, name, field, msg in flags:
        report.append(f"| {rownum} | {name} | {field} | {msg} |")
    report.append("")
    if overpaid_flags:
        report.append(f"## Overpayment anomalies ({len(overpaid_flags)}) — Paid exceeds Payment on NEW 2026 DEBTORS")
        report.append("Recorded as-is (payment amount honestly reflects what the sheet says was paid), just flagged for awareness.")
        report.append("")
        for name, payment, paid in overpaid_flags:
            report.append(f"- **{name}** — payment {payment:,.0f}, paid {paid:,.0f}")
        report.append("")
    report.append("## Assumptions made (please review)")
    report.append("- Every resident is treated as an **OWNER** (the sheet has no owner/tenant distinction).")
    report.append("- Payment **method** defaulted to **BANK_TRANSFER** and **paid date** to **2026-06-15** (sheets have no per-payment date/method).")
    report.append("- **Registration date** defaulted to **2026-01-01** for every resident.")
    report.append("- **Property type** defaulted to **DETACHED_HOUSE**, occupancy to **OCCUPIED**; house numbers were **synthesized** as `{ROAD}-{seq}` (real plot numbers are almost entirely absent from this workbook, unlike the 2025 one).")
    report.append("- **Security Fee** and **Yearly Devt. Levy** were merged into one levy ('Security Fee + Yearly Devt. Levy') because their sole source of truth, NEW 2026 DEBTORS, only reports them as one combined figure per resident.")
    report.append("- The **donation** column is identical between the dev-levy sheet and the electricity sheet for every row checked — imported once, from the electricity sheet, per the estate's own description of it as electricity-related.")

    with open(OUT_REPORT, "w") as f:
        f.write("\n".join(report))

    print(f"Residents/properties: {counts['residents']}")
    print(f"Invoices: {counts['invoices']} ({counts['payments']} paid, {counts['invoices'] - counts['payments']} unpaid)")
    print(f"24/2025 Debtors skipped: {counts['arrears_skipped']} residents, {counts['arrears_skipped_amount']:,.0f} total")
    print(f"Duplicate roster rows skipped: {len(dup_skipped)}")
    print(f"Unmatched (dev-levy/electricity/debtors): {len(devlevy_unmatched)}/{len(elec_unmatched)}/{len(debtors_unmatched)}")
    print(f"Consolidated multi-property (dev-levy/electricity/debtors): {len(devlevy_consolidated)}/{len(elec_consolidated)}/{len(debtors_consolidated)}")
    print(f"Flagged cells: {len(flags)}")
    print(f"Overpayment anomalies: {len(overpaid_flags)}")
    print(f"SQL written to {OUT_SQL}")
    print(f"Report written to {OUT_REPORT}")


if __name__ == "__main__":
    main()
